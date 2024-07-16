import xlwings as xw
import random
import openpyxl
dataframe = openpyxl.load_workbook("Book1.xlsx")
dataframe1 = dataframe.active
items = []
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        items.append(col[row].value)
del items[0:2]
new_items = []
f = 0
for i in items:
    f += 1
    if f % 2 == 1:
        new_items.append(i)
result1 = "B2:B" + str(len(new_items))
result2 = "A2:A" + str(len(new_items))


ws = xw.Book("Book1.xlsx").sheets['Sheet1']


english_words = ws.range(result2).value
your_languages_words = ws.range(result1).value

languages = ("YOUR LANGUAGE(EDIT)", "English")
good_answers = 0
def start():
    while True:
        try:
            a = str(input(f"Which language do you want to practice?{languages[0]}/{languages[1]}\n"))
            if a == languages[0]:
                main("1")
                break
            if a == languages[1]:
                main("2")
                break

        except (TypeError, ValueError):
            print("elírtad")

def main(n):
    global good_answers
    for i in range(len(english_words)):
        a = random.randrange(0, len(english_words))
        match n:
            case "1":
                print(good_answers)
                print(your_languages_words[a])
                e = input("")
                if e == english_words[a]:
                    print("good answer")
                    good_answers += 1
                else:
                    print("not right, but this is the proper answer: " + english_words[a] + "\n next:")
            case "2":
                print(good_answers)
                print(english_words[a])
                e = input("")
                if e == your_languages_words[a]:
                    print("good answer")
                    good_answers += 1
                else:
                    print("not right, but this is the proper answer: " + your_languages_words[a] + "\nnext:")
    print(f"here is your overall score: {good_answers}/{len(english_words)}")
    b = str(input("Do you want to try it again?(y/n)\n"))
    if b == "y":
        start()
    else:
        print("")

start()